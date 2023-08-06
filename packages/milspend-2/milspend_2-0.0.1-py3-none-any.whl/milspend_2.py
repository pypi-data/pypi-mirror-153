
import subprocess as sp
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import sys,os
if os.path.exists('./d2019_T5-01.xlsx'):
    wb=openpyxl.load_workbook('d2019_T5-01.xlsx')
else:
    sp.call("wget https://www.jil.go.jp/kokunai/statistics/databook/2019/05/d2019_T5-01.xlsx||wget https://github.com/NichihoYamauchi/milspend_2.py/blob/main/d2019_T5-01.xlsx",shell=True)
    wb=openpyxl.load_workbook('d2019_T5-01.xlsx')
#シート
sheet=wb['5-1']
#行を消す
sheet.delete_rows(1,4)
sheet.delete_rows(6,7)
sheet.delete_rows(14,40)

sheet.delete_cols(1)
sheet.delete_cols(4,10)
sheet.delete_cols(12,13)

#保存
wb.save('result.xlsx')
d=pd.read_excel('result.xlsx',engine='openpyxl',sheet_name='5-1')
size=0

x = ['日本', 'アメリカ', 'イギリス', 'ドイツ', 'フランス']
y = []

j = sheet.cell(9,11)
y.append(j)

a = sheet.cell(10,11)
y.append(a*145)

u = sheet.cell(11,11)
y.append(u*111)

g = sheet.cell(12,11)
y.append(g*166)

f = sheet.cell(13,11)
y.append(f*139)



plt.plot(x, y, marker="o")

def main():
    plt.legend()
    plt.savefig('result.png')
    plt.show()