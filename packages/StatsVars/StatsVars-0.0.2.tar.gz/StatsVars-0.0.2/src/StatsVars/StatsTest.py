#import math 
import numpy as np
import StatsVars as SVars


class Error(Exception):
    pass
class ConditionsNotMet(Error):
    pass

def parse_kwargs(keys, **kwargs):
    return SVars(parse_kwargs(keys, **kwargs))



rounding_point = 5
conditions_string = '''
    Are all conditions met?
        -Random
        -Less than 10% of population
        -Independent
        -Nearly Normal
            - Sample Size < 15: 
                ○ Only use t procedure if data are close to Normal
            - 15 < Sample Size < 50:
                ○ Only use if data is unimodal and reasonably symmetric
            - Sample Size > 40:
                ○ Use unless data is extremely skewed
    '''

def parse_nums(raw_values):
    str_values_list = raw_values.strip().split('\n')
    values = [float(i) for i in str_values_list]
    return values

def graph_histo(vals):
    from matplotlib import pyplot as plt    
    plt.hist(vals)
    plt.show()

def load_excel_data(location, col_needed, exceptions):
    from openpyxl import load_workbook

    spreadsheet = load_workbook(location)
    sheet_names = spreadsheet.sheetnames
    
    info_sheet = spreadsheet[sheet_names[0]]
    
    info = [row.value for row in info_sheet[col_needed] if row.value not in exceptions]
    return info


def two_sample_z_interval(t1, n1, t2, n2, c=.95):
    p1 = t1/n1
    p2 = t2/n2
    sd = SVars.two_sample_z_standard_deviation(p1, n1, p2, n2)
    difference = (p1 - p2)
    z_score = SVars.invNorm(c)
    
    total = z_score * sd
    interval = SVars.interval_notation(difference-total, difference+total, rounding=rounding_point)
    
    print(f'''
        (p1 - p2) +- z * sd
        ({p1} - {p2}) +- {z_score} * {sd}
        difference = {difference}
        z-score = {z_score}     sd = {sd}
        
        interval:
            {interval}
        '''
        )


def one_sample_t_interval_mean(vals = [], confidence = .95, **kwargs):
    '''
    Parameter
    Assumptions
    Name Test
    Interval
    Conclusion
    
    kwargs:
        -n: # of values
        -mu: data mean
        -sd: standard deviation 
    '''
    if not vals:
        keys = ['n', 'mu', 'sd']
        n, mu, sd = parse_kwargs(keys, **kwargs)
        degrees = n - 1
        SE = SVars.standard_error(sd=sd, n=n)
    else:
        n, mu, sd, SE = SVars.get_SVars(vals=vals, keys=['n', 'mu', 'sd', 'SE']).values()
        degrees = n - 1
    crit_score = SVars.invT(confidence=confidence, df=degrees)
    error = crit_score * SE

    interval = (mu - error, mu + error)
    #p = input('What are you investigating?')
    #graph_histo(values, SVars.count(vals))
    print(f'''
        crit score = {crit_score}
        mu: {mu}    sd: {sd}
        SE: {SE}    confidence: {confidence}
        {mu} +- {error} = {SVars.interval_notation(*interval, rounding=rounding_point)}
    ''')

def one_sample_t_test_mean(vals = [], expected_mean = 0, null = '=', hypo = '!=', alpha = .05, **kwargs):
    '''
        Parameter
        Hypothesis
        Assumptions
        Name test
        Test statistic
        Obtain p-value
        State conclusion
        
        **kwargs:
            -mu: expected mean
            -sd: standard deviation
            -n: # values
    '''
    if not vals:
        keys = ['mu', 'sd', 'n']
        mu, sd, n = SVars.parse_kwargs(keys, **kwargs)
        degrees = n - 1
        
    else:
        mu, sd, n = SVars.get_SVars(vals=vals, keys=['mu', 'sd', 'n']).values()
        degrees = n - 1
    t_score = SVars.t_score(expected_mean=expected_mean, mu=mu, sd=sd, n=n)
    p_score = 1 - SVars.Tcdf(t_score, degrees) if t_score > 0 else SVars.Tcdf(t_score, degrees)
    
    if hypo == '!=':
        p_score *= 2
    
    if p_score > alpha:
        #Accept null
        test_answer = f'{mu} (mu) {null} {expected_mean} (null)'
    else:
        #Reject null
        test_answer = f'{mu} (mu) {hypo} {expected_mean} (alt)'
    print(f'''
        null: mu {null} {expected_mean}
        alt: mu {hypo} {expected_mean}
        
        ans: {test_answer}
        
        mu: {mu}            sd: {sd}
        t-score: {t_score}     p-score: {p_score}
        df: {degrees}          n: {n}  
        '''
        )

def two_sample_t_interval(s1 = [], s2 = [], c = .95, **kwargs):
    '''
    kwargs:
        x1 = mean of s1
        x2 = mean of s2
        sd1 = standard deviation of s1
        sd2 = standard deviation of s2
        n1 = # of values of s1
        n2 = # of values of s2
    '''
    if not s1 and not s2:
        keys = ['x1', 'x2', 'sd1', 'sd2', 'n1', 'n2']
        x1, x2, sd1, sd2, n1, n2 = parse_kwargs(keys, **kwargs)
    else:
        x1, sd1, n1 = SVars.get_SVars(vals=s1, keys=['mu', 'sd', 'n']).values()
        x2, sd2, n2 = SVars.get_SVars(vals=s2, keys=['mu', 'sd', 'n']).values()
    SE = SVars.two_sample_error(sd1=sd1, sd2=sd2, n1=n1, n2=n2)
    df = SVars.two_sample_degrees_of_freedom(sd1=sd1, sd2=sd2, n1=n1, n2=n2)
    difference = x1 - x2
    t = SVars.invT(c, df)
    interval = (difference - (t * SE), difference + (t * SE))
    print(f'''
        {SVars.interval_notation(*interval, rounding=rounding_point)}
        x1: {x1}    x2: {x2}
        sd1: {sd1}  sd2: {sd2}
        n1: {n1}    n2: {n2}
        SE: {SE}    df: {df}
        t: {t}  
        ''')

def two_sample_t_test(s1 = [], s2 = [], null = '=', hypo = '!=', alpha = .05, **kwargs):
    '''kwargs:
        x1 = mean of s1
        x2 = mean of s2
        sd1 = standard deviation of s1
        sd2 = standard deviation of s2
        n1 = # of values of s1
        n2 = # of values of s2
    '''
    if not s1 and not s2:
        keys = ['x1', 'x2', 'sd1', 'sd2', 'n1', 'n2']
        x1, x2, sd1, sd2, n1, n2 = parse_kwargs(keys, **kwargs)
    else:
        x1, sd1, n1 = SVars.get_SVars(vals=s1, keys=['mu', 'sd', 'n']).values()
        x2, sd2, n2 = SVars.get_SVars(vals=s2, keys=['mu', 'sd', 'n']).values()
    SE = SVars.two_sample_error(sd1=sd1, sd2=sd2, n1=n1, n2=n2)
    df = SVars.two_sample_degrees_of_freedom(sd1=sd1, sd2=sd2, n1=n1, n2=n2) 
    t_score = SVars.two_sample_t_score(x1=x1, sd1=sd1, n1=n1, x2=x2, sd2=sd2, n2=n2)
    p_score = 1 - SVars.Tcdf(t_score, df) if t_score > 0 else SVars.Tcdf(t_score, df)
    
    if hypo == '!=':
        p_score *= 2
    
    if p_score > alpha:
        #Accept null
        test_answer = f's1 {null} s2'
    else:
        #Reject null
        test_answer = f's1 {hypo} s2'
    print(f'''  
        {test_answer}
        x1: {x1}    x2: {x2}
        sd1: {sd1}  sd2: {sd2}
        n1: {n1}    n2: {n2}
        SE: {SE}    df: {df}
        t score: {t_score}  p score: {p_score}
        alpha: {alpha}
        ''')
    #summary


def paired_t_test(s1 = [], s2 = [], expected_difference = 0, null = '=', hypo = '!=', alpha = .05, **kwargs):
    '''
    kwargs:
        -mu: mean of differences
        -sd: standard deviation of differences
        -n: # of pairs of differences
    '''
    if not s1 and not s2:
        keys = ['mu', 'sd', 'n']
        mu, sd, n = parse_kwargs(keys, **kwargs)
    else:
        mu, sd, n = SVars.get_SVars(SVars.differences(s1, s2), ['mu', 'sd', 'n']).values()
    df = n - 1
    t_score = SVars.pairwise_t_score(mu=mu, sd=sd, n=n)
    p_score = 1 - SVars.Tcdf(t_score, df) if t_score > 0 else SVars.Tcdf(t_score, df)    
    
    if hypo == '!=':
        p_score *= 2
    
    if p_score > alpha:
        #Accept null
        test_answer = f'mean difference {null} {expected_difference}'
    else:
        #Reject null
        test_answer = f'mean difference {hypo} {expected_difference}'
    print(f'''  
        null: mean difference {null} {expected_difference}
        alt: mean difference {hypo} {expected_difference}
        
        ans: {test_answer}
            s1:                s2:
        mu: {SVars.mean(s1)}             {SVars.mean(s2)}
        sd: {SVars.sample_standard_deviation(s1)}  {SVars.sample_standard_deviation(s2)}
        n: {SVars.count(s1)}             {SVars.count(s2)}
        t score: {t_score}  p score: {p_score}
        alpha: {alpha}
        ''')

def paired_t_interval(s1 = [], s2 = [], c = .95, **kwargs):
    '''
    kwargs:
        -mu: mean of differences
        -sd: standard deviation of differences
        -n: # of pairs of differences
    '''
    if not s1 and not s2:
        keys = ['mu', 'sd', 'n']
        mu, sd, n = SVars.parse_kwargs(keys, **kwargs)
    else:
        mu, sd, n = SVars.get_SVars(SVars.differences(s1, s2), ['mu', 'sd', 'n']).values()
    print(mu, sd, n)
    df = n-1
    t = SVars.invT(c, df)
    SE = SVars.standard_error(sd=sd, n=n)
    interval = (mu - (t*SE), mu + (t*SE))
    
    if s1 and s2:
        print(f'''
            {SVars.interval_notation(*interval, rounding=rounding_point)}
            s1:                s2:
            mu: {SVars.mean(s1)}             {SVars.mean(s2)}
            sd: {SVars.sample_standard_deviation(s1)}  {SVars.sample_standard_deviation(s2)}
            n: {SVars.count(s1)}             {SVars.count(s2)}
            confidence: {c} t: {t}
            SE: {SE}
        ''')
    else:
        print(f'''
            {SVars.interval_notation(*interval, rounding=rounding_point)}
            mu: {mu}    sd: {sd}
            n:  {n}     df: {df}
            confidence: {c} t: {t}
            SE: {SE}
        ''')


def goodness_fit_test(counts, expected, null = '=', hypo = '!=', alpha = .05):
    '''
    Conditions:
        Counted Data - data are counts of categorical SVars
        Randomization Condition
        At least 5 counts in each cell
    '''
    n = SVars.count(counts)
    df = n - 1
    chi_score = SVars.X_score(counts, expected)
    p_score = SVars.Xcdf(chi_score, df)
    
    crit_val = SVars.X_crit_value(df=df, alpha=alpha)
    
    if chi_score > crit_val:
        #reject null
        test_ans = f'{chi_score} > {crit_val} Reject Null'
    else:
        #dont reject null
        test_ans = f'{chi_score} < {crit_val} Accept Null'
    print(f'''
        {counts} - {expected}
        n: {n}    df: {df}
        X^2 = {chi_score}     p: {p_score}
        crit = {crit_val}
        {test_ans}
        ''')

def independence_test(categories_true, categories_false, alpha = .05, cols = 2):
    categories_table = ''
    for i in range(len(categories_true)):
        categories_table += str(categories_true[i]) + '\t' + str(categories_false[i]) + '\n'
    categories_table.strip('\n')
    
    expected_true, expected_false = SVars.X_expected(categories_true, categories_false, rounding_point)
    expectations_table = ''
    for i in range(len(expected_true)):
        expectations_table += str(expected_true[i]) + '\t' + str(expected_false[i]) + '\n'
    expectations_table.strip('\n')
    
    chi_score = SVars.X_score(vals=list(categories_true + categories_false), expected=list(expected_true + expected_false))
    df = len(expected_true) - 1 * (cols - 1)
    p_score = SVars.Xcdf(chi_score, df)

    if p_score >= alpha:
        #accept null
        test_ans = categories_table, '=', expectations_table, 'Accept Null'
    else:
        #reject null
        test_ans = categories_table, '!=', expectations_table, 'Reject Null'

    print(f'''
        ans: 
        {test_ans[0]} 
        {test_ans[1]} 
        {test_ans[2]}
        n: {df + 1}     df: {df}
        X^2: {chi_score}    p: {p_score}
        {test_ans[-1]}
        ''')


def regression_line(x_vals = [], y_vals = [], c=.95, x_title = 'x-axis', y_title = 'y-axis', B0 = 0, B1 = 0, hypo ='!=', graph=False, prediction=None):
    from matplotlib import pyplot as plt
    '''
        y = B0 + B1x + e
        1. Linearity Assumption
        2. Independece Assumption
        3. Equal Variance Assumption
        4. Normal Population Assumption
    '''
    n = SVars.count(x_vals)
    df = n - 2
    
    b0 = SVars.linreg_intercept(x_vals=x_vals, y_vals=y_vals)
    b1 = SVars.linreg_slope(x_vals=x_vals, y_vals=y_vals)
    R = SVars.correlation_coefficient(x_vals=x_vals, y_vals=y_vals)
    
    eq = f'{b0} + {b1} * x'
    mse = SVars.mean_squared_error(expected=y_vals, predicted=SVars.get_cords(x_vals=x_vals, line=eq)) #approx sd
    
    crit_score = SVars.invT(c, df)

    residuals = SVars.residuals(x_vals=x_vals, y_vals=y_vals, eq=eq)
    res_sd = SVars.residual_standard_deviation(x_vals=x_vals, y_vals=y_vals, eq=eq)
    
    slope_SE = SVars.slope_standard_error(res_sd=res_sd, x_values=x_vals)
    slope_t_ratio = SVars.regression_t_score(b=b1, B=B1, SE=slope_SE)
    slope_p_val = 1 - SVars.Tcdf(slope_t_ratio, df) if slope_t_ratio > 0 else SVars.Tcdf(slope_t_ratio, df)

    int_SE = SVars.intercept_standard_error(mse, x_vals)
    int_t_ratio = SVars.regression_t_score(b=b0, B=B0, SE=int_SE)
    int_p_val = 1 - SVars.Tcdf(int_t_ratio, df) if int_t_ratio > 0 else SVars.Tcdf(int_t_ratio, df)
    
    if hypo == '!=':
        slope_p_val *= 2
        int_p_val *= 2
        
    slope_sd = crit_score * slope_SE
    int_sd = crit_score * int_SE
    if prediction:
        p = eval(eq, {'x' : prediction})
        margin = SVars.prediction_interval_error(crit_score, slope_SE, prediction, SVars.mean(x_vals), res_sd, n)
        prediction_int = SVars.interval_notation(*(p-margin, p+margin), rounding=rounding_point)
    else:
        prediction_int = None
    #report
    print(f'''
        Dependant Variable: {y_title}
        Independent Variable: {x_title}
        
        n: {n}    df: {df}
        {y_title} = {b0} {SVars.sign(b1)} {abs(b1)}({x_title})
        Sample Size: {SVars.count(x_vals)}
        R (correlation coefficient) = {R}
        R-sq = {R**2}
        
        residual sd = {res_sd}
        crit score = {crit_score}

        {x_title}:
            t ratio: {slope_t_ratio}
            p value: {slope_p_val}

            mean x = {SVars.mean(x_vals)}
            SE(b1) = {slope_SE}
            interval:
                stat += {slope_sd} (slope variability: {SVars.interval_notation(*(b1-slope_sd, b1+slope_sd), rounding=rounding_point)})
        
        {y_title}:
            t ratio: {int_t_ratio}
            p_value: {int_p_val}
            
            mean_y = {SVars.mean(y_vals)}
            SE(b0) = {int_SE}
            interval:
                stat += {int_sd} (intercept variability: {SVars.interval_notation(*(b0-int_sd, b0+int_sd), rounding=rounding_point)})
        
        
    '''  
        )
    if graph:
        plt.title('Plotted Points/Line')
        plt.plot(x_vals, y_vals, 'o')
        x = np.linspace(*SVars.min_max(x_vals), 50)
        plt.plot(x, SVars.get_cords(x, eq))
        plt.xlabel(x_title)
        plt.ylabel(y_title)
        plt.show()
        
        plt.plot(x_vals, residuals, 'ro')
        plt.title('Residuals Plot')
        plt.show()

if __name__ == '__main__':
    #sample use
    vaccine_n = 14_134 
    vaccine_covid = 11

    placebo_n = 14_073
    placebo_covid = 185
    two_sample_z_interval(vaccine_covid, vaccine_n, placebo_covid, placebo_n)
    regression_line([2, 5, 7, 9], [5, 3, 1, -2], c=.95, graph=True, prediction=None)
